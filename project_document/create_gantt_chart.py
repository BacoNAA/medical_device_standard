#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
医疗器械标准数据库与应用系统 - 甘特图生成脚本
基于 WBS v2.0 和 PDM网络图 v1.0
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

def create_gantt_chart():
    """创建详细的项目甘特图"""
    
    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "项目甘特图"
    
    # 项目时间范围
    project_start = datetime(2025, 10, 13)
    project_end = datetime(2025, 12, 19)
    
    # 定义颜色方案（简洁专业）
    colors = {
        'header': '4472C4',           # 深蓝色 - 表头
        'phase': '305496',            # 深蓝 - 阶段
        'module': '5B9BD5',           # 中蓝 - 模块
        'subfunction': 'DDEBF7',      # 浅蓝 - 子功能
        'planning': 'E7E6E6',         # 浅灰 - 规划管理
        'design': 'E7E6E6',           # 浅灰 - 设计
        'frontend': 'E7E6E6',         # 浅灰 - 前端开发
        'backend': 'E7E6E6',          # 浅灰 - 后端开发
        'testing': 'E7E6E6',          # 浅灰 - 测试
        'deployment': 'E7E6E6',       # 浅灰 - 部署
        'bar_phase': '305496',        # 深蓝 - 阶段条
        'bar_module': '5B9BD5',       # 中蓝 - 模块条
        'bar_task': '9DC3E6',         # 浅蓝 - 任务条
    }
    
    # 责任小组时间条颜色方案（增强对比度，更易区分）
    resource_colors = {
        '项目经理': '808080',                           # 深灰色 - 项目管理
        'UI/UX组': 'FF9900',                            # 鲜橙色 - UI设计
        '前端组': '4472C4',                             # 深蓝色 - 前端开发
        '后端组': '548235',                             # 深绿色 - 后端开发
        '测试组': 'FFD700',                             # 金黄色 - 测试
        '全体': '7030A0',                               # 深紫色 - 全体协作
        # 组合责任方（多团队协作）
        'UI/UX组+前端组+后端组+测试组': 'C55A11',      # 深橙红色 - 全功能团队协作
        '项目经理+后端组+UI/UX组': '5B9BD5',           # 中蓝色 - 管理+技术
        '后端组+测试组': '70AD47',                      # 亮绿色 - 后端+测试
        '前端组+后端组': '00B0F0',                      # 天蓝色 - 前后端协作
        '后端组+测试组与部署': '92D050',                # 浅绿色 - 后端+测试+部署
        '测试组与部署': 'FFC000',                       # 亮黄色 - 测试与部署
        '测试组与部署+前端组': '8FAADC',                # 浅蓝紫 - 测试+前端
    }
    
    # 设置列宽
    ws.column_dimensions['A'].width = 8   # WBS编号
    ws.column_dimensions['B'].width = 35  # 任务名称
    ws.column_dimensions['C'].width = 12  # 开始日期
    ws.column_dimensions['D'].width = 12  # 结束日期
    ws.column_dimensions['E'].width = 8   # 工期
    ws.column_dimensions['F'].width = 15  # 责任方
    
    # 日期列从G列开始
    date_start_col = 7  # G列
    
    # 创建日期列（按周显示）
    current_date = project_start
    col_idx = date_start_col
    date_columns = {}  # 日期到列号的映射
    
    while current_date <= project_end:
        week_num = (current_date - project_start).days // 7 + 1
        date_str = current_date.strftime('%m/%d')
        ws.cell(row=1, column=col_idx, value=f"W{week_num}\n{date_str}")
        ws.column_dimensions[get_column_letter(col_idx)].width = 3.5
        
        # 记录日期映射
        date_columns[current_date.strftime('%Y-%m-%d')] = col_idx
        
        current_date += timedelta(days=1)
        col_idx += 1
    
    # 设置表头样式
    header_fill = PatternFill(start_color=colors['header'], end_color=colors['header'], fill_type='solid')
    header_font = Font(bold=True, size=10)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # 设置固定列表头
    headers = ['WBS编号', '任务名称', '开始日期', '结束日期', '工期(天)', '责任方']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
    
    # 设置日期列表头样式
    for col_idx in range(date_start_col, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = Font(bold=True, size=8)
        cell.alignment = center_align
    
    # 定义项目数据结构
    tasks = []
    
    # ==================== 阶段1：基础设施建设 ====================
    tasks.append({
        'wbs': '阶段1', 'name': '【阶段1】基础设施建设', 
        'start': '2025-10-13', 'end': '2025-11-16', 'duration': 35,
        'resource': '全体', 'type': 'phase'
    })
    
    # 模块1：项目管理与基础设施
    tasks.append({
        'wbs': '1.0', 'name': '模块1：项目管理与基础设施', 
        'start': '2025-10-13', 'end': '2025-11-16', 'duration': 35,
        'resource': '项目经理+后端组+UI/UX组', 'type': 'module'
    })
    
    tasks.append({
        'wbs': '1.1', 'name': '  项目规划与管理', 
        'start': '2025-10-13', 'end': '2025-10-26', 'duration': 14,
        'resource': '项目经理', 'type': 'planning'
    })
    tasks.append({
        'wbs': '1.1.1', 'name': '    项目启动', 
        'start': '2025-10-13', 'end': '2025-10-19', 'duration': 7,
        'resource': '项目经理', 'type': 'planning'
    })
    tasks.append({
        'wbs': '1.1.2', 'name': '    资源与风险管理', 
        'start': '2025-10-20', 'end': '2025-10-26', 'duration': 7,
        'resource': '项目经理', 'type': 'planning'
    })
    
    tasks.append({
        'wbs': '1.2', 'name': '  需求分析', 
        'start': '2025-10-13', 'end': '2025-10-26', 'duration': 14,
        'resource': '项目经理', 'type': 'planning'
    })
    tasks.append({
        'wbs': '1.2.1', 'name': '    需求调研与分析', 
        'start': '2025-10-13', 'end': '2025-10-19', 'duration': 7,
        'resource': '项目经理', 'type': 'planning'
    })
    tasks.append({
        'wbs': '1.2.2', 'name': '    需求规格说明书', 
        'start': '2025-10-20', 'end': '2025-10-26', 'duration': 7,
        'resource': '项目经理', 'type': 'planning'
    })
    
    tasks.append({
        'wbs': '1.3', 'name': '  系统架构设计', 
        'start': '2025-10-27', 'end': '2025-11-02', 'duration': 7,
        'resource': '后端组', 'type': 'design'
    })
    tasks.append({
        'wbs': '1.3.1', 'name': '    架构与技术选型', 
        'start': '2025-10-27', 'end': '2025-10-30', 'duration': 4,
        'resource': '后端组', 'type': 'design'
    })
    tasks.append({
        'wbs': '1.3.2', 'name': '    系统设计文档', 
        'start': '2025-10-31', 'end': '2025-11-02', 'duration': 3,
        'resource': '后端组', 'type': 'design'
    })
    
    tasks.append({
        'wbs': '1.4', 'name': '  数据库设计与实现', 
        'start': '2025-11-03', 'end': '2025-11-16', 'duration': 14,
        'resource': '后端组', 'type': 'backend'
    })
    tasks.append({
        'wbs': '1.4.1', 'name': '    数据库设计', 
        'start': '2025-11-03', 'end': '2025-11-09', 'duration': 7,
        'resource': '后端组', 'type': 'backend'
    })
    tasks.append({
        'wbs': '1.4.2', 'name': '    数据库实施', 
        'start': '2025-11-10', 'end': '2025-11-16', 'duration': 7,
        'resource': '后端组', 'type': 'backend'
    })
    
    tasks.append({
        'wbs': '1.5', 'name': '  UI/UX整体设计', 
        'start': '2025-11-03', 'end': '2025-11-16', 'duration': 14,
        'resource': 'UI/UX组', 'type': 'design'
    })
    tasks.append({
        'wbs': '1.5.1', 'name': '    界面设计', 
        'start': '2025-11-03', 'end': '2025-11-09', 'duration': 7,
        'resource': 'UI/UX组', 'type': 'design'
    })
    tasks.append({
        'wbs': '1.5.2', 'name': '    设计规范文档', 
        'start': '2025-11-10', 'end': '2025-11-16', 'duration': 7,
        'resource': 'UI/UX组', 'type': 'design'
    })
    
    # ==================== 阶段2：基础功能模块 ====================
    tasks.append({
        'wbs': '阶段2', 'name': '【阶段2】基础功能模块开发', 
        'start': '2025-11-17', 'end': '2025-11-30', 'duration': 14,
        'resource': '全体', 'type': 'phase'
    })
    
    # 模块2：用户管理功能
    tasks.append({
        'wbs': '2.0', 'name': '模块2：用户管理功能', 
        'start': '2025-11-17', 'end': '2025-11-30', 'duration': 14,
        'resource': 'UI/UX组+前端组+后端组+测试组', 'type': 'module'
    })
    
    tasks.append({
        'wbs': '2.1', 'name': '  用户管理UI设计', 
        'start': '2025-11-17', 'end': '2025-11-18', 'duration': 2,
        'resource': 'UI/UX组', 'type': 'design'
    })
    tasks.append({
        'wbs': '2.1.1', 'name': '    用户管理界面设计', 
        'start': '2025-11-17', 'end': '2025-11-18', 'duration': 2,
        'resource': 'UI/UX组', 'type': 'design'
    })
    
    tasks.append({
        'wbs': '2.2', 'name': '  用户登录/注册前端开发', 
        'start': '2025-11-19', 'end': '2025-11-25', 'duration': 7,
        'resource': '前端组', 'type': 'frontend'
    })
    tasks.append({
        'wbs': '2.2.1', 'name': '    登录页面开发', 
        'start': '2025-11-19', 'end': '2025-11-19', 'duration': 1,
        'resource': '前端组', 'type': 'frontend'
    })
    tasks.append({
        'wbs': '2.2.2', 'name': '    记住密码功能', 
        'start': '2025-11-20', 'end': '2025-11-20', 'duration': 1,
        'resource': '前端组', 'type': 'frontend'
    })
    tasks.append({
        'wbs': '2.2.3', 'name': '    验证码集成', 
        'start': '2025-11-21', 'end': '2025-11-21', 'duration': 1,
        'resource': '前端组', 'type': 'frontend'
    })
    tasks.append({
        'wbs': '2.2.4', 'name': '    用户创建界面', 
        'start': '2025-11-22', 'end': '2025-11-22', 'duration': 1,
        'resource': '前端组', 'type': 'frontend'
    })
    tasks.append({
        'wbs': '2.2.5', 'name': '    用户信息编辑界面', 
        'start': '2025-11-23', 'end': '2025-11-24', 'duration': 2,
        'resource': '前端组', 'type': 'frontend'
    })
    tasks.append({
        'wbs': '2.2.6', 'name': '    密码修改界面', 
        'start': '2025-11-25', 'end': '2025-11-25', 'duration': 1,
        'resource': '前端组', 'type': 'frontend'
    })
    
    tasks.append({
        'wbs': '2.3', 'name': '  用户认证授权后端开发', 
        'start': '2025-11-19', 'end': '2025-11-25', 'duration': 7,
        'resource': '后端组', 'type': 'backend'
    })
    tasks.append({
        'wbs': '2.3.1', 'name': '    登录验证API', 
        'start': '2025-11-19', 'end': '2025-11-19', 'duration': 1,
        'resource': '后端组', 'type': 'backend'
    })
    tasks.append({
        'wbs': '2.3.2', 'name': '    密码加密存储', 
        'start': '2025-11-20', 'end': '2025-11-20', 'duration': 1,
        'resource': '后端组', 'type': 'backend'
    })
    tasks.append({
        'wbs': '2.3.3', 'name': '    会话管理', 
        'start': '2025-11-21', 'end': '2025-11-21', 'duration': 1,
        'resource': '后端组', 'type': 'backend'
    })
    tasks.append({
        'wbs': '2.3.4', 'name': '    角色自动识别', 
        'start': '2025-11-22', 'end': '2025-11-22', 'duration': 1,
        'resource': '后端组', 'type': 'backend'
    })
    tasks.append({
        'wbs': '2.3.5', 'name': '    权限验证机制', 
        'start': '2025-11-23', 'end': '2025-11-24', 'duration': 2,
        'resource': '后端组', 'type': 'backend'
    })
    tasks.append({
        'wbs': '2.3.6', 'name': '    用户管理API', 
        'start': '2025-11-25', 'end': '2025-11-25', 'duration': 1,
        'resource': '后端组', 'type': 'backend'
    })
    
    tasks.append({
        'wbs': '2.4', 'name': '  用户管理功能测试', 
        'start': '2025-11-26', 'end': '2025-11-28', 'duration': 3,
        'resource': '质量管理和测试组', 'type': 'testing'
    })
    tasks.append({
        'wbs': '2.4.1', 'name': '    登录功能测试', 
        'start': '2025-11-26', 'end': '2025-11-26', 'duration': 1,
        'resource': '质量管理和测试组', 'type': 'testing'
    })
    tasks.append({
        'wbs': '2.4.2', 'name': '    用户管理功能测试', 
        'start': '2025-11-27', 'end': '2025-11-27', 'duration': 1,
        'resource': '质量管理和测试组', 'type': 'testing'
    })
    tasks.append({
        'wbs': '2.4.3', 'name': '    权限控制测试', 
        'start': '2025-11-28', 'end': '2025-11-28', 'duration': 1,
        'resource': '质量管理和测试组', 'type': 'testing'
    })
    
    tasks.append({
        'wbs': '2.5', 'name': '  用户管理功能部署', 
        'start': '2025-11-29', 'end': '2025-11-30', 'duration': 2,
        'resource': '后端组', 'type': 'deployment'
    })
    tasks.append({
        'wbs': '2.5.1', 'name': '    用户模块部署', 
        'start': '2025-11-29', 'end': '2025-11-30', 'duration': 2,
        'resource': '后端组', 'type': 'deployment'
    })
    
    # 模块3：标准数据管理功能（并行开发）
    tasks.append({
        'wbs': '3.0', 'name': '模块3：标准数据管理功能', 
        'start': '2025-11-17', 'end': '2025-11-30', 'duration': 14,
        'resource': 'UI/UX组+前端组+后端组+测试组', 'type': 'module'
    })
    
    tasks.append({
        'wbs': '3.1', 'name': '  标准数据管理UI设计', 
        'start': '2025-11-17', 'end': '2025-11-18', 'duration': 2,
        'resource': 'UI/UX组', 'type': 'design'
    })
    tasks.append({
        'wbs': '3.1.1', 'name': '    标准管理界面设计', 
        'start': '2025-11-17', 'end': '2025-11-18', 'duration': 2,
        'resource': 'UI/UX组', 'type': 'design'
    })
    
    tasks.append({
        'wbs': '3.2', 'name': '  标准浏览与检索前端开发', 
        'start': '2025-11-19', 'end': '2025-11-25', 'duration': 7,
        'resource': '前端组', 'type': 'frontend'
    })
    tasks.append({
        'wbs': '3.2.1', 'name': '    标准列表页面', 
        'start': '2025-11-19', 'end': '2025-11-19', 'duration': 1,
        'resource': '前端组', 'type': 'frontend'
    })
    tasks.append({
        'wbs': '3.2.2', 'name': '    标准详情页面', 
        'start': '2025-11-20', 'end': '2025-11-20', 'duration': 1,
        'resource': '前端组', 'type': 'frontend'
    })
    tasks.append({
        'wbs': '3.2.3', 'name': '    检验项目列表', 
        'start': '2025-11-21', 'end': '2025-11-21', 'duration': 1,
        'resource': '前端组', 'type': 'frontend'
    })
    tasks.append({
        'wbs': '3.2.4', 'name': '    说明性内容展示', 
        'start': '2025-11-22', 'end': '2025-11-22', 'duration': 1,
        'resource': '前端组', 'type': 'frontend'
    })
    tasks.append({
        'wbs': '3.2.5', 'name': '    附录内容展示', 
        'start': '2025-11-23', 'end': '2025-11-23', 'duration': 1,
        'resource': '前端组', 'type': 'frontend'
    })
    tasks.append({
        'wbs': '3.2.6', 'name': '    多条件检索界面', 
        'start': '2025-11-24', 'end': '2025-11-24', 'duration': 1,
        'resource': '前端组', 'type': 'frontend'
    })
    tasks.append({
        'wbs': '3.2.7', 'name': '    关键词高亮显示', 
        'start': '2025-11-25', 'end': '2025-11-25', 'duration': 1,
        'resource': '前端组', 'type': 'frontend'
    })
    
    tasks.append({
        'wbs': '3.3', 'name': '  Excel/Word导入解析后端开发', 
        'start': '2025-11-19', 'end': '2025-11-23', 'duration': 5,
        'resource': '后端组', 'type': 'backend'
    })
    tasks.append({
        'wbs': '3.3.1', 'name': '    文件上传API', 
        'start': '2025-11-19', 'end': '2025-11-19', 'duration': 1,
        'resource': '后端组', 'type': 'backend'
    })
    tasks.append({
        'wbs': '3.3.2', 'name': '    Excel解析引擎', 
        'start': '2025-11-19', 'end': '2025-11-19', 'duration': 1,
        'resource': '后端组', 'type': 'backend'
    })
    tasks.append({
        'wbs': '3.3.3', 'name': '    附录智能链接', 
        'start': '2025-11-20', 'end': '2025-11-20', 'duration': 1,
        'resource': '后端组', 'type': 'backend'
    })
    tasks.append({
        'wbs': '3.3.4', 'name': '    设备智能识别', 
        'start': '2025-11-21', 'end': '2025-11-21', 'duration': 1,
        'resource': '后端组', 'type': 'backend'
    })
    tasks.append({
        'wbs': '3.3.5', 'name': '    数据预览生成', 
        'start': '2025-11-22', 'end': '2025-11-22', 'duration': 1,
        'resource': '后端组', 'type': 'backend'
    })
    tasks.append({
        'wbs': '3.3.6', 'name': '    批量数据入库', 
        'start': '2025-11-22', 'end': '2025-11-22', 'duration': 1,
        'resource': '后端组', 'type': 'backend'
    })
    tasks.append({
        'wbs': '3.3.7', 'name': '    Word解析引擎', 
        'start': '2025-11-23', 'end': '2025-11-23', 'duration': 1,
        'resource': '后端组', 'type': 'backend'
    })
    
    tasks.append({
        'wbs': '3.4', 'name': '  标准数据存储与查询后端开发', 
        'start': '2025-11-24', 'end': '2025-11-25', 'duration': 2,
        'resource': '后端组', 'type': 'backend'
    })
    tasks.append({
        'wbs': '3.4.1', 'name': '    标准查询API', 
        'start': '2025-11-24', 'end': '2025-11-24', 'duration': 1,
        'resource': '后端组', 'type': 'backend'
    })
    tasks.append({
        'wbs': '3.4.2', 'name': '    全文检索功能', 
        'start': '2025-11-24', 'end': '2025-11-24', 'duration': 1,
        'resource': '后端组', 'type': 'backend'
    })
    tasks.append({
        'wbs': '3.4.3', 'name': '    搜索结果排序', 
        'start': '2025-11-25', 'end': '2025-11-25', 'duration': 1,
        'resource': '后端组', 'type': 'backend'
    })
    tasks.append({
        'wbs': '3.4.4', 'name': '    标准信息编辑API', 
        'start': '2025-11-25', 'end': '2025-11-25', 'duration': 1,
        'resource': '后端组', 'type': 'backend'
    })
    tasks.append({
        'wbs': '3.4.5', 'name': '    检验项目编辑API', 
        'start': '2025-11-25', 'end': '2025-11-25', 'duration': 1,
        'resource': '后端组', 'type': 'backend'
    })
    tasks.append({
        'wbs': '3.4.6', 'name': '    数据删除API', 
        'start': '2025-11-25', 'end': '2025-11-25', 'duration': 1,
        'resource': '后端组', 'type': 'backend'
    })
    
    tasks.append({
        'wbs': '3.5', 'name': '  标准数据管理功能测试', 
        'start': '2025-11-26', 'end': '2025-11-28', 'duration': 3,
        'resource': '质量管理和测试组', 'type': 'testing'
    })
    tasks.append({
        'wbs': '3.5.1', 'name': '    导入功能测试', 
        'start': '2025-11-26', 'end': '2025-11-26', 'duration': 1,
        'resource': '质量管理和测试组', 'type': 'testing'
    })
    tasks.append({
        'wbs': '3.5.2', 'name': '    浏览检索测试', 
        'start': '2025-11-27', 'end': '2025-11-27', 'duration': 1,
        'resource': '质量管理和测试组', 'type': 'testing'
    })
    tasks.append({
        'wbs': '3.5.3', 'name': '    数据维护测试', 
        'start': '2025-11-28', 'end': '2025-11-28', 'duration': 1,
        'resource': '质量管理和测试组', 'type': 'testing'
    })
    
    tasks.append({
        'wbs': '3.6', 'name': '  标准数据管理功能部署', 
        'start': '2025-11-29', 'end': '2025-11-30', 'duration': 2,
        'resource': '后端组', 'type': 'deployment'
    })
    tasks.append({
        'wbs': '3.6.1', 'name': '    标准数据模块部署', 
        'start': '2025-11-29', 'end': '2025-11-30', 'duration': 2,
        'resource': '后端组', 'type': 'deployment'
    })
    
    # 模块4：设备管理功能（并行开发）
    tasks.append({
        'wbs': '4.0', 'name': '模块4：设备管理功能', 
        'start': '2025-11-17', 'end': '2025-11-30', 'duration': 14,
        'resource': 'UI/UX组+前端组+后端组+测试组', 'type': 'module'
    })
    
    tasks.append({
        'wbs': '4.1', 'name': '  设备管理UI设计', 
        'start': '2025-11-17', 'end': '2025-11-18', 'duration': 2,
        'resource': 'UI/UX组', 'type': 'design'
    })
    tasks.append({
        'wbs': '4.1.1', 'name': '    设备管理界面设计', 
        'start': '2025-11-17', 'end': '2025-11-18', 'duration': 2,
        'resource': 'UI/UX组', 'type': 'design'
    })
    
    tasks.append({
        'wbs': '4.2', 'name': '  设备管理前端开发', 
        'start': '2025-11-19', 'end': '2025-11-25', 'duration': 7,
        'resource': '前端组', 'type': 'frontend'
    })
    tasks.append({
        'wbs': '4.2.1', 'name': '    设备列表页面', 
        'start': '2025-11-19', 'end': '2025-11-19', 'duration': 1,
        'resource': '前端组', 'type': 'frontend'
    })
    tasks.append({
        'wbs': '4.2.2', 'name': '    设备信息录入界面', 
        'start': '2025-11-20', 'end': '2025-11-21', 'duration': 2,
        'resource': '前端组', 'type': 'frontend'
    })
    tasks.append({
        'wbs': '4.2.3', 'name': '    设备信息编辑界面', 
        'start': '2025-11-22', 'end': '2025-11-22', 'duration': 1,
        'resource': '前端组', 'type': 'frontend'
    })
    tasks.append({
        'wbs': '4.2.4', 'name': '    设备关联界面', 
        'start': '2025-11-23', 'end': '2025-11-24', 'duration': 2,
        'resource': '前端组', 'type': 'frontend'
    })
    tasks.append({
        'wbs': '4.2.5', 'name': '    关键词库维护界面', 
        'start': '2025-11-25', 'end': '2025-11-25', 'duration': 1,
        'resource': '前端组', 'type': 'frontend'
    })
    
    tasks.append({
        'wbs': '4.3', 'name': '  设备管理后端开发', 
        'start': '2025-11-19', 'end': '2025-11-25', 'duration': 7,
        'resource': '后端组', 'type': 'backend'
    })
    tasks.append({
        'wbs': '4.3.1', 'name': '    设备编号生成', 
        'start': '2025-11-19', 'end': '2025-11-19', 'duration': 1,
        'resource': '后端组', 'type': 'backend'
    })
    tasks.append({
        'wbs': '4.3.2', 'name': '    设备CRUD API', 
        'start': '2025-11-20', 'end': '2025-11-21', 'duration': 2,
        'resource': '后端组', 'type': 'backend'
    })
    tasks.append({
        'wbs': '4.3.3', 'name': '    自动关联功能', 
        'start': '2025-11-22', 'end': '2025-11-22', 'duration': 1,
        'resource': '后端组', 'type': 'backend'
    })
    tasks.append({
        'wbs': '4.3.4', 'name': '    手动关联API', 
        'start': '2025-11-23', 'end': '2025-11-23', 'duration': 1,
        'resource': '后端组', 'type': 'backend'
    })
    tasks.append({
        'wbs': '4.3.5', 'name': '    关键词匹配算法', 
        'start': '2025-11-24', 'end': '2025-11-24', 'duration': 1,
        'resource': '后端组', 'type': 'backend'
    })
    tasks.append({
        'wbs': '4.3.6', 'name': '    关键词库管理API', 
        'start': '2025-11-25', 'end': '2025-11-25', 'duration': 1,
        'resource': '后端组', 'type': 'backend'
    })
    
    tasks.append({
        'wbs': '4.4', 'name': '  设备管理功能测试', 
        'start': '2025-11-26', 'end': '2025-11-28', 'duration': 3,
        'resource': '质量管理和测试组', 'type': 'testing'
    })
    tasks.append({
        'wbs': '4.4.1', 'name': '    设备信息管理测试', 
        'start': '2025-11-26', 'end': '2025-11-27', 'duration': 2,
        'resource': '质量管理和测试组', 'type': 'testing'
    })
    tasks.append({
        'wbs': '4.4.2', 'name': '    设备关联功能测试', 
        'start': '2025-11-28', 'end': '2025-11-28', 'duration': 1,
        'resource': '质量管理和测试组', 'type': 'testing'
    })
    
    tasks.append({
        'wbs': '4.5', 'name': '  设备管理功能部署', 
        'start': '2025-11-29', 'end': '2025-11-30', 'duration': 2,
        'resource': '后端组', 'type': 'deployment'
    })
    tasks.append({
        'wbs': '4.5.1', 'name': '    设备模块部署', 
        'start': '2025-11-29', 'end': '2025-11-30', 'duration': 2,
        'resource': '后端组', 'type': 'deployment'
    })
    
    # ==================== 阶段3：业务功能模块 ====================
    tasks.append({
        'wbs': '阶段3', 'name': '【阶段3】业务功能模块开发', 
        'start': '2025-12-01', 'end': '2025-12-14', 'duration': 14,
        'resource': '全体', 'type': 'phase'
    })
    
    # 模块5：错误报告功能
    tasks.append({
        'wbs': '5.0', 'name': '模块5：错误报告功能', 
        'start': '2025-12-01', 'end': '2025-12-14', 'duration': 14,
        'resource': 'UI/UX组+前端组+后端组+测试组', 'type': 'module'
    })
    
    tasks.append({
        'wbs': '5.1', 'name': '  错误报告UI设计', 
        'start': '2025-12-01', 'end': '2025-12-02', 'duration': 2,
        'resource': 'UI/UX组', 'type': 'design'
    })
    tasks.append({
        'wbs': '5.1.1', 'name': '    报告管理界面设计', 
        'start': '2025-12-01', 'end': '2025-12-02', 'duration': 2,
        'resource': 'UI/UX组', 'type': 'design'
    })
    
    tasks.append({
        'wbs': '5.2', 'name': '  错误报告前端开发', 
        'start': '2025-12-03', 'end': '2025-12-09', 'duration': 7,
        'resource': '前端组', 'type': 'frontend'
    })
    tasks.append({
        'wbs': '5.2.1', 'name': '    报告提交表单', 
        'start': '2025-12-03', 'end': '2025-12-04', 'duration': 2,
        'resource': '前端组', 'type': 'frontend'
    })
    tasks.append({
        'wbs': '5.2.2', 'name': '    报告列表页面', 
        'start': '2025-12-05', 'end': '2025-12-06', 'duration': 2,
        'resource': '前端组', 'type': 'frontend'
    })
    tasks.append({
        'wbs': '5.2.3', 'name': '    报告详情页面', 
        'start': '2025-12-07', 'end': '2025-12-08', 'duration': 2,
        'resource': '前端组', 'type': 'frontend'
    })
    tasks.append({
        'wbs': '5.2.4', 'name': '    处理操作界面', 
        'start': '2025-12-09', 'end': '2025-12-09', 'duration': 1,
        'resource': '前端组', 'type': 'frontend'
    })
    
    tasks.append({
        'wbs': '5.3', 'name': '  错误报告后端开发', 
        'start': '2025-12-03', 'end': '2025-12-09', 'duration': 7,
        'resource': '后端组', 'type': 'backend'
    })
    tasks.append({
        'wbs': '5.3.1', 'name': '    报告编号生成', 
        'start': '2025-12-03', 'end': '2025-12-03', 'duration': 1,
        'resource': '后端组', 'type': 'backend'
    })
    tasks.append({
        'wbs': '5.3.2', 'name': '    报告提交API', 
        'start': '2025-12-04', 'end': '2025-12-05', 'duration': 2,
        'resource': '后端组', 'type': 'backend'
    })
    tasks.append({
        'wbs': '5.3.3', 'name': '    报告查询API', 
        'start': '2025-12-06', 'end': '2025-12-07', 'duration': 2,
        'resource': '后端组', 'type': 'backend'
    })
    tasks.append({
        'wbs': '5.3.4', 'name': '    报告处理API', 
        'start': '2025-12-08', 'end': '2025-12-08', 'duration': 1,
        'resource': '后端组', 'type': 'backend'
    })
    tasks.append({
        'wbs': '5.3.5', 'name': '    通知功能', 
        'start': '2025-12-09', 'end': '2025-12-09', 'duration': 1,
        'resource': '后端组', 'type': 'backend'
    })
    
    tasks.append({
        'wbs': '5.4', 'name': '  错误报告功能测试', 
        'start': '2025-12-10', 'end': '2025-12-12', 'duration': 3,
        'resource': '质量管理和测试组', 'type': 'testing'
    })
    tasks.append({
        'wbs': '5.4.1', 'name': '    报告提交测试', 
        'start': '2025-12-10', 'end': '2025-12-11', 'duration': 2,
        'resource': '质量管理和测试组', 'type': 'testing'
    })
    tasks.append({
        'wbs': '5.4.2', 'name': '    报告处理测试', 
        'start': '2025-12-12', 'end': '2025-12-12', 'duration': 1,
        'resource': '质量管理和测试组', 'type': 'testing'
    })
    
    tasks.append({
        'wbs': '5.5', 'name': '  错误报告功能部署', 
        'start': '2025-12-13', 'end': '2025-12-14', 'duration': 2,
        'resource': '后端组', 'type': 'deployment'
    })
    tasks.append({
        'wbs': '5.5.1', 'name': '    错误报告模块部署', 
        'start': '2025-12-13', 'end': '2025-12-14', 'duration': 2,
        'resource': '后端组', 'type': 'deployment'
    })
    
    # 模块6：数据导出功能（并行开发）
    tasks.append({
        'wbs': '6.0', 'name': '模块6：数据导出功能', 
        'start': '2025-12-01', 'end': '2025-12-14', 'duration': 14,
        'resource': 'UI/UX组+前端组+后端组+测试组', 'type': 'module'
    })
    
    tasks.append({
        'wbs': '6.1', 'name': '  数据导出UI设计', 
        'start': '2025-12-01', 'end': '2025-12-02', 'duration': 2,
        'resource': 'UI/UX组', 'type': 'design'
    })
    tasks.append({
        'wbs': '6.1.1', 'name': '    导出功能界面设计', 
        'start': '2025-12-01', 'end': '2025-12-02', 'duration': 2,
        'resource': 'UI/UX组', 'type': 'design'
    })
    
    tasks.append({
        'wbs': '6.2', 'name': '  导出配置界面前端开发', 
        'start': '2025-12-03', 'end': '2025-12-09', 'duration': 7,
        'resource': '前端组', 'type': 'frontend'
    })
    tasks.append({
        'wbs': '6.2.1', 'name': '    表1导出界面', 
        'start': '2025-12-03', 'end': '2025-12-03', 'duration': 1,
        'resource': '前端组', 'type': 'frontend'
    })
    tasks.append({
        'wbs': '6.2.2', 'name': '    表2导出界面', 
        'start': '2025-12-04', 'end': '2025-12-04', 'duration': 1,
        'resource': '前端组', 'type': 'frontend'
    })
    tasks.append({
        'wbs': '6.2.3', 'name': '    表3导出界面', 
        'start': '2025-12-05', 'end': '2025-12-05', 'duration': 1,
        'resource': '前端组', 'type': 'frontend'
    })
    tasks.append({
        'wbs': '6.2.4', 'name': '    导出预览组件', 
        'start': '2025-12-06', 'end': '2025-12-06', 'duration': 1,
        'resource': '前端组', 'type': 'frontend'
    })
    tasks.append({
        'wbs': '6.2.5', 'name': '    模板创建界面', 
        'start': '2025-12-07', 'end': '2025-12-07', 'duration': 1,
        'resource': '前端组', 'type': 'frontend'
    })
    tasks.append({
        'wbs': '6.2.6', 'name': '    字段配置组件', 
        'start': '2025-12-08', 'end': '2025-12-08', 'duration': 1,
        'resource': '前端组', 'type': 'frontend'
    })
    tasks.append({
        'wbs': '6.2.7', 'name': '    模板列表页面', 
        'start': '2025-12-09', 'end': '2025-12-09', 'duration': 1,
        'resource': '前端组', 'type': 'frontend'
    })
    
    tasks.append({
        'wbs': '6.3', 'name': '  表1/表2/表3导出后端开发', 
        'start': '2025-12-03', 'end': '2025-12-06', 'duration': 4,
        'resource': '后端组', 'type': 'backend'
    })
    tasks.append({
        'wbs': '6.3.1', 'name': '    表1导出API', 
        'start': '2025-12-03', 'end': '2025-12-03', 'duration': 1,
        'resource': '后端组', 'type': 'backend'
    })
    tasks.append({
        'wbs': '6.3.2', 'name': '    表2导出API', 
        'start': '2025-12-04', 'end': '2025-12-04', 'duration': 1,
        'resource': '后端组', 'type': 'backend'
    })
    tasks.append({
        'wbs': '6.3.3', 'name': '    表3导出API', 
        'start': '2025-12-05', 'end': '2025-12-05', 'duration': 1,
        'resource': '后端组', 'type': 'backend'
    })
    tasks.append({
        'wbs': '6.3.4', 'name': '    Excel生成模块', 
        'start': '2025-12-06', 'end': '2025-12-06', 'duration': 1,
        'resource': '后端组', 'type': 'backend'
    })
    
    tasks.append({
        'wbs': '6.4', 'name': '  自定义导出后端开发', 
        'start': '2025-12-07', 'end': '2025-12-09', 'duration': 3,
        'resource': '后端组', 'type': 'backend'
    })
    tasks.append({
        'wbs': '6.4.1', 'name': '    模板保存API', 
        'start': '2025-12-07', 'end': '2025-12-07', 'duration': 1,
        'resource': '后端组', 'type': 'backend'
    })
    tasks.append({
        'wbs': '6.4.2', 'name': '    模板分享功能', 
        'start': '2025-12-08', 'end': '2025-12-08', 'duration': 1,
        'resource': '后端组', 'type': 'backend'
    })
    tasks.append({
        'wbs': '6.4.3', 'name': '    模板查询API', 
        'start': '2025-12-08', 'end': '2025-12-08', 'duration': 1,
        'resource': '后端组', 'type': 'backend'
    })
    tasks.append({
        'wbs': '6.4.4', 'name': '    自定义导出API', 
        'start': '2025-12-09', 'end': '2025-12-09', 'duration': 1,
        'resource': '后端组', 'type': 'backend'
    })
    
    tasks.append({
        'wbs': '6.5', 'name': '  数据导出功能测试', 
        'start': '2025-12-10', 'end': '2025-12-12', 'duration': 3,
        'resource': '质量管理和测试组', 'type': 'testing'
    })
    tasks.append({
        'wbs': '6.5.1', 'name': '    预设模板导出测试', 
        'start': '2025-12-10', 'end': '2025-12-11', 'duration': 2,
        'resource': '质量管理和测试组', 'type': 'testing'
    })
    tasks.append({
        'wbs': '6.5.2', 'name': '    自定义导出测试', 
        'start': '2025-12-12', 'end': '2025-12-12', 'duration': 1,
        'resource': '质量管理和测试组', 'type': 'testing'
    })
    
    tasks.append({
        'wbs': '6.6', 'name': '  数据导出功能部署', 
        'start': '2025-12-13', 'end': '2025-12-14', 'duration': 2,
        'resource': '后端组', 'type': 'deployment'
    })
    tasks.append({
        'wbs': '6.6.1', 'name': '    数据导出模块部署', 
        'start': '2025-12-13', 'end': '2025-12-14', 'duration': 2,
        'resource': '后端组', 'type': 'deployment'
    })
    
    # 模块7：PTR编辑功能（并行开发）
    tasks.append({
        'wbs': '7.0', 'name': '模块7：PTR编辑功能', 
        'start': '2025-12-01', 'end': '2025-12-14', 'duration': 14,
        'resource': 'UI/UX组+前端组+后端组+测试组', 'type': 'module'
    })
    
    tasks.append({
        'wbs': '7.1', 'name': '  PTR编辑器UI设计', 
        'start': '2025-12-01', 'end': '2025-12-02', 'duration': 2,
        'resource': 'UI/UX组', 'type': 'design'
    })
    tasks.append({
        'wbs': '7.1.1', 'name': '    PTR编辑器界面设计', 
        'start': '2025-12-01', 'end': '2025-12-02', 'duration': 2,
        'resource': 'UI/UX组', 'type': 'design'
    })
    
    tasks.append({
        'wbs': '7.2', 'name': '  富文本编辑器前端开发', 
        'start': '2025-12-03', 'end': '2025-12-06', 'duration': 4,
        'resource': '前端组', 'type': 'frontend'
    })
    tasks.append({
        'wbs': '7.2.1', 'name': '    文档创建界面', 
        'start': '2025-12-03', 'end': '2025-12-03', 'duration': 1,
        'resource': '前端组', 'type': 'frontend'
    })
    tasks.append({
        'wbs': '7.2.2', 'name': '    文档列表页面', 
        'start': '2025-12-03', 'end': '2025-12-03', 'duration': 1,
        'resource': '前端组', 'type': 'frontend'
    })
    tasks.append({
        'wbs': '7.2.3', 'name': '    富文本编辑器集成', 
        'start': '2025-12-04', 'end': '2025-12-04', 'duration': 1,
        'resource': '前端组', 'type': 'frontend'
    })
    tasks.append({
        'wbs': '7.2.4', 'name': '    章节树导航', 
        'start': '2025-12-05', 'end': '2025-12-05', 'duration': 1,
        'resource': '前端组', 'type': 'frontend'
    })
    tasks.append({
        'wbs': '7.2.5', 'name': '    自动保存功能', 
        'start': '2025-12-06', 'end': '2025-12-06', 'duration': 1,
        'resource': '前端组', 'type': 'frontend'
    })
    tasks.append({
        'wbs': '7.2.6', 'name': '    字数统计功能', 
        'start': '2025-12-06', 'end': '2025-12-06', 'duration': 1,
        'resource': '前端组', 'type': 'frontend'
    })
    
    tasks.append({
        'wbs': '7.3', 'name': '  智能辅助编辑前端开发', 
        'start': '2025-12-07', 'end': '2025-12-09', 'duration': 3,
        'resource': '前端组', 'type': 'frontend'
    })
    tasks.append({
        'wbs': '7.3.1', 'name': '    模板插入组件', 
        'start': '2025-12-07', 'end': '2025-12-07', 'duration': 1,
        'resource': '前端组', 'type': 'frontend'
    })
    tasks.append({
        'wbs': '7.3.2', 'name': '    关键词联想查找', 
        'start': '2025-12-08', 'end': '2025-12-08', 'duration': 1,
        'resource': '前端组', 'type': 'frontend'
    })
    tasks.append({
        'wbs': '7.3.3', 'name': '    标准内容插入', 
        'start': '2025-12-08', 'end': '2025-12-08', 'duration': 1,
        'resource': '前端组', 'type': 'frontend'
    })
    tasks.append({
        'wbs': '7.3.4', 'name': '    智能辅助面板', 
        'start': '2025-12-09', 'end': '2025-12-09', 'duration': 1,
        'resource': '前端组', 'type': 'frontend'
    })
    
    tasks.append({
        'wbs': '7.4', 'name': '  PTR存储与Word生成后端开发', 
        'start': '2025-12-03', 'end': '2025-12-09', 'duration': 7,
        'resource': '后端组', 'type': 'backend'
    })
    tasks.append({
        'wbs': '7.4.1', 'name': '    文档管理API', 
        'start': '2025-12-03', 'end': '2025-12-03', 'duration': 1,
        'resource': '后端组', 'type': 'backend'
    })
    tasks.append({
        'wbs': '7.4.2', 'name': '    文档权限控制', 
        'start': '2025-12-04', 'end': '2025-12-04', 'duration': 1,
        'resource': '后端组', 'type': 'backend'
    })
    tasks.append({
        'wbs': '7.4.3', 'name': '    内容保存API', 
        'start': '2025-12-05', 'end': '2025-12-05', 'duration': 1,
        'resource': '后端组', 'type': 'backend'
    })
    tasks.append({
        'wbs': '7.4.4', 'name': '    联想查找API', 
        'start': '2025-12-06', 'end': '2025-12-06', 'duration': 1,
        'resource': '后端组', 'type': 'backend'
    })
    tasks.append({
        'wbs': '7.4.5', 'name': '    第3章框架生成', 
        'start': '2025-12-07', 'end': '2025-12-07', 'duration': 1,
        'resource': '后端组', 'type': 'backend'
    })
    tasks.append({
        'wbs': '7.4.6', 'name': '    Word格式转换', 
        'start': '2025-12-08', 'end': '2025-12-08', 'duration': 1,
        'resource': '后端组', 'type': 'backend'
    })
    tasks.append({
        'wbs': '7.4.7', 'name': '    格式规范处理', 
        'start': '2025-12-08', 'end': '2025-12-08', 'duration': 1,
        'resource': '后端组', 'type': 'backend'
    })
    tasks.append({
        'wbs': '7.4.8', 'name': '    文件生成与下载', 
        'start': '2025-12-09', 'end': '2025-12-09', 'duration': 1,
        'resource': '后端组', 'type': 'backend'
    })
    
    tasks.append({
        'wbs': '7.5', 'name': '  PTR编辑功能测试', 
        'start': '2025-12-10', 'end': '2025-12-12', 'duration': 3,
        'resource': '质量管理和测试组', 'type': 'testing'
    })
    tasks.append({
        'wbs': '7.5.1', 'name': '    编辑器功能测试', 
        'start': '2025-12-10', 'end': '2025-12-10', 'duration': 1,
        'resource': '质量管理和测试组', 'type': 'testing'
    })
    tasks.append({
        'wbs': '7.5.2', 'name': '    智能辅助测试', 
        'start': '2025-12-11', 'end': '2025-12-11', 'duration': 1,
        'resource': '质量管理和测试组', 'type': 'testing'
    })
    tasks.append({
        'wbs': '7.5.3', 'name': '    Word导出测试', 
        'start': '2025-12-12', 'end': '2025-12-12', 'duration': 1,
        'resource': '质量管理和测试组', 'type': 'testing'
    })
    
    tasks.append({
        'wbs': '7.6', 'name': '  PTR编辑功能部署', 
        'start': '2025-12-13', 'end': '2025-12-14', 'duration': 2,
        'resource': '后端组', 'type': 'deployment'
    })
    tasks.append({
        'wbs': '7.6.1', 'name': '    PTR编辑模块部署', 
        'start': '2025-12-13', 'end': '2025-12-14', 'duration': 2,
        'resource': '后端组', 'type': 'deployment'
    })
    
    # ==================== 阶段4：系统测试与交付 ====================
    tasks.append({
        'wbs': '阶段4', 'name': '【阶段4】系统测试与交付', 
        'start': '2025-12-15', 'end': '2025-12-19', 'duration': 5,
        'resource': '全体', 'type': 'phase'
    })
    
    # 模块8：系统测试与交付
    tasks.append({
        'wbs': '8.0', 'name': '模块8：系统测试与交付', 
        'start': '2025-12-15', 'end': '2025-12-19', 'duration': 5,
        'resource': '全体', 'type': 'module'
    })
    
    tasks.append({
        'wbs': '8.1', 'name': '  系统集成测试', 
        'start': '2025-12-15', 'end': '2025-12-15', 'duration': 1,
        'resource': '质量管理和测试组', 'type': 'testing'
    })
    tasks.append({
        'wbs': '8.1.1', 'name': '    集成测试', 
        'start': '2025-12-15', 'end': '2025-12-15', 'duration': 1,
        'resource': '质量管理和测试组', 'type': 'testing'
    })
    
    tasks.append({
        'wbs': '8.2', 'name': '  性能与安全测试', 
        'start': '2025-12-15', 'end': '2025-12-15', 'duration': 1,
        'resource': '质量管理和测试组', 'type': 'testing'
    })
    tasks.append({
        'wbs': '8.2.1', 'name': '    综合测试', 
        'start': '2025-12-15', 'end': '2025-12-15', 'duration': 1,
        'resource': '质量管理和测试组', 'type': 'testing'
    })
    
    tasks.append({
        'wbs': '8.3', 'name': '  用户验收测试(UAT)', 
        'start': '2025-12-16', 'end': '2025-12-18', 'duration': 3,
        'resource': '质量管理和测试组+前端组', 'type': 'testing'
    })
    tasks.append({
        'wbs': '8.3.1', 'name': '    UAT准备与执行', 
        'start': '2025-12-16', 'end': '2025-12-17', 'duration': 2,
        'resource': '质量管理和测试组', 'type': 'testing'
    })
    tasks.append({
        'wbs': '8.3.2', 'name': '    问题修复与验收', 
        'start': '2025-12-18', 'end': '2025-12-18', 'duration': 1,
        'resource': '前端组', 'type': 'testing'
    })
    
    tasks.append({
        'wbs': '8.4', 'name': '  系统部署', 
        'start': '2025-12-19', 'end': '2025-12-19', 'duration': 1,
        'resource': '后端组+质量管理和测试组', 'type': 'deployment'
    })
    tasks.append({
        'wbs': '8.4.1', 'name': '    生产环境部署', 
        'start': '2025-12-19', 'end': '2025-12-19', 'duration': 1,
        'resource': '后端组', 'type': 'deployment'
    })
    tasks.append({
        'wbs': '8.4.2', 'name': '    部署验证', 
        'start': '2025-12-19', 'end': '2025-12-19', 'duration': 1,
        'resource': '质量管理和测试组', 'type': 'deployment'
    })
    
    tasks.append({
        'wbs': '8.5', 'name': '  项目验收与收尾', 
        'start': '2025-12-19', 'end': '2025-12-19', 'duration': 1,
        'resource': '项目经理', 'type': 'deployment'
    })
    tasks.append({
        'wbs': '8.5.1', 'name': '    项目验收', 
        'start': '2025-12-19', 'end': '2025-12-19', 'duration': 1,
        'resource': '项目经理', 'type': 'deployment'
    })
    tasks.append({
        'wbs': '8.5.2', 'name': '    项目总结', 
        'start': '2025-12-19', 'end': '2025-12-19', 'duration': 1,
        'resource': '项目经理', 'type': 'deployment'
    })
    
    # 填充任务数据
    row = 2
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    for task in tasks:
        # WBS编号
        cell_wbs = ws.cell(row=row, column=1, value=task['wbs'])
        cell_wbs.alignment = Alignment(horizontal='left', vertical='center')
        cell_wbs.border = thin_border
        
        # 任务名称
        cell_name = ws.cell(row=row, column=2, value=task['name'])
        cell_name.alignment = Alignment(horizontal='left', vertical='center')
        cell_name.border = thin_border
        
        # 开始日期
        cell_start = ws.cell(row=row, column=3, value=task['start'])
        cell_start.alignment = Alignment(horizontal='center', vertical='center')
        cell_start.border = thin_border
        
        # 结束日期
        cell_end = ws.cell(row=row, column=4, value=task['end'])
        cell_end.alignment = Alignment(horizontal='center', vertical='center')
        cell_end.border = thin_border
        
        # 工期
        cell_duration = ws.cell(row=row, column=5, value=task['duration'])
        cell_duration.alignment = Alignment(horizontal='center', vertical='center')
        cell_duration.border = thin_border
        
        # 责任方
        cell_resource = ws.cell(row=row, column=6, value=task['resource'])
        cell_resource.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell_resource.border = thin_border
        
        # 设置行样式（左侧列）
        task_type = task['type']
        if task_type == 'phase':
            fill_color = colors['phase']
            font_style = Font(bold=True, size=11, color='FFFFFF')
        elif task_type == 'module':
            fill_color = colors['module']
            font_style = Font(bold=True, size=10, color='FFFFFF')
        else:  # 所有子任务统一样式
            fill_color = colors['planning']
            font_style = Font(size=9)
        
        # 根据责任方确定时间条颜色
        resource = task['resource']
        bar_color = resource_colors.get(resource, '9DC3E6')  # 默认浅蓝色
        
        # 应用样式到前6列
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
            cell.font = font_style
        
        # 绘制甘特图条
        start_date = datetime.strptime(task['start'], '%Y-%m-%d')
        end_date = datetime.strptime(task['end'], '%Y-%m-%d')
        
        current = start_date
        while current <= end_date:
            date_key = current.strftime('%Y-%m-%d')
            if date_key in date_columns:
                col_idx = date_columns[date_key]
                cell = ws.cell(row=row, column=col_idx)
                cell.fill = PatternFill(start_color=bar_color, end_color=bar_color, fill_type='solid')
                cell.border = thin_border
            current += timedelta(days=1)
        
        row += 1
    
    # 添加图例
    legend_row = row + 2
    ws.cell(row=legend_row, column=1, value='图例说明：').font = Font(bold=True, size=11)
    
    # 左侧任务行颜色图例
    legend_items_left = [
        ('阶段', colors['phase'], '项目主要阶段'),
        ('模块', colors['module'], '功能模块'),
        ('子任务', colors['planning'], '具体工作任务'),
    ]
    
    legend_row += 1
    ws.cell(row=legend_row, column=1, value='任务行颜色：').font = Font(bold=True, size=10)
    legend_row += 1
    
    for idx, (name, color, desc) in enumerate(legend_items_left):
        col_offset = (idx % 3) * 3 + 1
        row_offset = legend_row + (idx // 3)
        
        # 颜色块
        cell_color = ws.cell(row=row_offset, column=col_offset)
        cell_color.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        cell_color.border = thin_border
        
        # 说明文字
        cell_text = ws.cell(row=row_offset, column=col_offset + 1, value=f'{name} - {desc}')
        cell_text.alignment = Alignment(horizontal='left', vertical='center')
        cell_text.font = Font(size=9)
    
    # 右侧时间条颜色图例
    legend_row += 2
    ws.cell(row=legend_row, column=1, value='时间条颜色（按责任小组）：').font = Font(bold=True, size=10)
    legend_row += 1
    
    legend_items_bar = [
        ('项目经理', resource_colors['项目经理'], '项目管理工作'),
        ('UI/UX组', resource_colors['UI/UX组'], 'UI设计工作'),
        ('前端组', resource_colors['前端组'], '前端开发工作'),
        ('后端组', resource_colors['后端组'], '后端开发工作'),
        ('测试组', resource_colors['测试组'], '测试工作'),
        ('全体', resource_colors['全体'], '全体协作'),
        ('多团队协作', 'B4A7D6', '多个团队共同工作'),
    ]
    
    for idx, (name, color, desc) in enumerate(legend_items_bar):
        col_offset = (idx % 3) * 3 + 1
        row_offset = legend_row + (idx // 3)
        
        # 颜色块
        cell_color = ws.cell(row=row_offset, column=col_offset)
        cell_color.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        cell_color.border = thin_border
        
        # 说明文字
        cell_text = ws.cell(row=row_offset, column=col_offset + 1, value=f'{name} - {desc}')
        cell_text.alignment = Alignment(horizontal='left', vertical='center')
        cell_text.font = Font(size=9)
    
    # 冻结窗格
    ws.freeze_panes = 'G2'
    
    # 设置行高
    ws.row_dimensions[1].height = 30
    for r in range(2, row):
        ws.row_dimensions[r].height = 20
    
    # 保存文件
    filename = r'D:\软件项目管理\project_document\项目甘特图 v1.0.xlsx'
    wb.save(filename)
    print(f'[OK] 甘特图已生成：{filename}')
    print(f'[OK] 包含 {len(tasks)} 个任务')
    print(f'[OK] 项目周期：{project_start.strftime("%Y-%m-%d")} 至 {project_end.strftime("%Y-%m-%d")}')
    print(f'[OK] 总工期：{(project_end - project_start).days + 1} 天')

if __name__ == '__main__':
    create_gantt_chart()

